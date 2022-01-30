import pandas as pd
import openpyxl
import sys
import tkinter as tk
from tkinter import filedialog
MANUAALI = False
VERBOSE = False


if len(sys.argv) > 1:
    try:
        input_file = str(sys.argv[1])
        panda_data = pd.read_csv(input_file, header=None)
    except:
        print("Could not open file. Second argument should be input_file in form of xxxx.csv")
        print(str(sys.argv))
        quit()
else:
    try:
        root = tk.Tk()
        root.withdraw()
        print("Choose your file, only use pdf-files")
        input_file = filedialog.askopenfilenames()
        input_file = input_file[0]
        panda_data = pd.read_csv(input_file, header=None)
    except ValueError:
        print("Check file name, quitting")
        quit(0)
    except KeyboardInterrupt:
        print("Keyboard Interrupt, quitting")
        quit(0)
    except FileNotFoundError:
        print("File not found, quitting")
        quit(0)

if MANUAALI:
    input_mitattavia_reagensseja = int(input("Number of measured compounds: "))
else:
    input_mitattavia_reagensseja = int((len(panda_data.columns)-7)/4)

mitattavia_reagensseja = [2]
for i in range(input_mitattavia_reagensseja):
    mitattavia_reagensseja.append((i+2)*4)
    mitattavia_reagensseja.append((i+2)*4+1)

panda_data.iloc[0, :] = panda_data.iloc[0, :].shift()
panda_stripped = panda_data[mitattavia_reagensseja]

panda_stripped.columns = list(panda_stripped.loc[0])
panda_stripped = panda_stripped.drop(0)


if VERBOSE:
    print(panda_stripped)

output_file = input_file[:-4]+"_output.xlsx"
print(f"Saving file to: {output_file}\n")

panda_stripped.to_excel(output_file, sheet_name='output')

# Reopening for openpyxl 
commas_pois = openpyxl.load_workbook(output_file)
ws = commas_pois["output"]

i = 0
for r in range(3,ws.max_row+1):
    for c in range(3,ws.max_column+1):
        s = ws.cell(r,c).value
        if s != None:
            ws.cell(r,c).value = float(ws.cell(r,c).value)
            i += 1

commas_pois.save(output_file)

print("{} cells fixed. \nReady".format(i))
input("Press Enter to continue.")
